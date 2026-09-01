"""Generovanie XLSX výkazu (jeden hárok, bez osôb) — príloha k VYÚČTOVANIU (#13).

`build_xlsx(od, do, items, client_name) -> bytes` postaví workbook presne podľa
formátovej špecifikácie vlastníka: jeden hárok, titulok s obdobím, hlavička,
dátové riadky (Dátum / Hodiny / Popis činnosti) so zebrou a riadok SPOLU.
Mená osôb sa do súboru NIKDY nezapisujú — vlastnícke rozhodnutie „jeden súbor,
jeden hárok, žiadni ľudia".
"""
import io
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Farby ako ARGB (alfa FF) — presne ako vo vzorke vlastníka.
_TITLE_FILL = PatternFill('solid', fgColor='FF1F4E78')
_HEADER_FILL = PatternFill('solid', fgColor='FF2E75B6')
_ZEBRA_FILL = PatternFill('solid', fgColor='FFF2F2F2')
_TOTAL_FILL = PatternFill('solid', fgColor='FFDDEBF7')
_WHITE = 'FFFFFFFF'

_CENTER = Alignment(horizontal='center', vertical='center')
_LEFT = Alignment(horizontal='left', vertical='center')

# Zdieľateľný font dátových buniek (openpyxl štýly sú immutable/zdieľateľné) —
# jedna inštancia namiesto novej pre každú bunku.
_DATA_FONT = Font(size=10)

_DATE_FMT = r'yyyy\-mm\-dd'
_HOURS_FMT = '0.0'

# Znaky zakázané v názve hárku Excelu ([ ] : * ? / \) + limit 31 znakov.
_FORBIDDEN_SHEET = set('[]:*?/') | {chr(92)}


def _sanitize_sheet_name(raw):
    """Bezpečný názov hárku: zakázané znaky ([ ] : * ? / \\) → medzera,
    orez apostrofov na okrajoch, max 31 znakov (limit Excelu)."""
    name = ''.join(' ' if ch in _FORBIDDEN_SHEET else ch for ch in raw)
    return name.strip("'")[:31]


def _sheet_name(od, do):
    return _sanitize_sheet_name(f'{od:%d.%m.} – {do:%d.%m.%Y}')


def _ascii_slug(s):
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    return re.sub(r'[^A-Za-z0-9]+', '_', s).strip('_')


def xlsx_filename(od, do, client_name):
    """ASCII-safe názov súboru: Vykaz_prace_<od>_<do>[_<Klient>].xlsx."""
    base = f'Vykaz_prace_{od:%Y-%m-%d}_{do:%Y-%m-%d}'
    if client_name:
        slug = _ascii_slug(client_name)
        if slug:
            base += f'_{slug}'
    return base + '.xlsx'


def build_xlsx(od, do, items, client_name):
    """Postav XLSX výkaz a vráť ho ako bytes.

    ``items`` = položky settlementu ``[(dátum, autor, hodiny, popis), ...]``
    chronologicky. ``autor`` sa IGNORUJE — do súboru sa mená nedávajú.
    ``client_name`` None/prázdne → riadok Klient sa vynechá (layout ostáva).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_name(od, do)

    # Titulok — merged A1:C1
    ws.merge_cells('A1:C1')
    a1 = ws['A1']
    a1.value = f'Výkaz práce – {od:%d.%m.%Y} → {do:%d.%m.%Y}'
    for coord in ('A1', 'B1', 'C1'):
        ws[coord].fill = _TITLE_FILL
    a1.font = Font(bold=True, size=14, color=_WHITE)
    a1.alignment = _CENTER

    # Klient — merged A2:C2 (len ak je nastavený)
    if client_name:
        ws.merge_cells('A2:C2')
        a2 = ws['A2']
        a2.value = f'Klient: {client_name}'
        a2.font = Font(size=10)
        a2.alignment = _LEFT

    # riadok 3 prázdny; hlavička v riadku 4
    for coord, label in (('A4', 'Dátum'), ('B4', 'Hodiny'), ('C4', 'Popis činnosti')):
        c = ws[coord]
        c.value = label
        c.font = Font(bold=True, color=_WHITE)
        c.fill = _HEADER_FILL
        c.alignment = _CENTER

    # dátové riadky od 5
    first = 5
    row = first
    for i, (date, _author, hours, desc) in enumerate(items):
        zebra = _ZEBRA_FILL if i % 2 == 1 else None
        a = ws.cell(row=row, column=1, value=date.date())
        a.number_format = _DATE_FMT
        a.alignment = _CENTER
        a.font = _DATA_FONT
        b = ws.cell(row=row, column=2, value=hours)
        b.number_format = _HOURS_FMT
        b.alignment = _CENTER
        b.font = _DATA_FONT
        c = ws.cell(row=row, column=3, value=desc)
        c.alignment = _LEFT
        c.font = _DATA_FONT
        # Ochrana proti injektáži vzorca: openpyxl bunku so stringom začínajúcim
        # '=' uloží ako FORMULU (data_type 'f'). Popis pochádza zo správ kanála,
        # takže „= …" by v klientskom súbore ožil ako vzorec — vynúť text.
        if isinstance(desc, str) and desc.startswith('='):
            c.data_type = 's'
        if zebra:
            for cell in (a, b, c):
                cell.fill = zebra
        row += 1

    last = row - 1  # posledný dátový riadok (= 4 keď žiadne položky)
    total_row = row
    ws.cell(row=total_row, column=1, value='SPOLU')
    b = ws.cell(row=total_row, column=2,
                value=f'=SUM(B{first}:B{last})' if items else 0)
    b.number_format = _HOURS_FMT
    b.alignment = _CENTER
    ws.cell(row=total_row, column=3, value='hodín')
    for col in (1, 2, 3):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = _TOTAL_FILL
        cell.font = Font(bold=True, size=10)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
