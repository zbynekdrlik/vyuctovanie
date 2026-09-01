"""Testy generovania XLSX výkazu (jeden hárok, bez osôb) — #13.

Fiktívny klient a fiktívne popisy — žiadne reálne zákaznícke údaje.
"""
import datetime as dt
import io

import openpyxl

from vyuct.xlsx import build_xlsx, xlsx_filename, _sanitize_sheet_name

TZ = dt.timezone(dt.timedelta(hours=2))
OD = dt.datetime(2026, 8, 19, 10, 0, tzinfo=TZ)
DO = dt.datetime(2026, 9, 1, 18, 0, tzinfo=TZ)

ITEMS = [
    (dt.datetime(2026, 8, 19, 10, 0, tzinfo=TZ), 'Ján Novák', 3.0, 'návrh výkresu'),
    (dt.datetime(2026, 8, 20, 9, 0, tzinfo=TZ), 'Peter Kováč', 6.0, 'oprava modelu'),
    (dt.datetime(2026, 8, 25, 14, 0, tzinfo=TZ), 'Ján Novák', 1.5, 'deploy'),
]


def _load(items=ITEMS, client='Testovací klient', od=OD, do=DO):
    data = build_xlsx(od, do, items, client)
    assert isinstance(data, (bytes, bytearray))
    return openpyxl.load_workbook(io.BytesIO(data))


def _as_date(v):
    return v.date() if isinstance(v, dt.datetime) else v


def test_returns_single_sheet_named_by_period():
    wb = _load()
    assert len(wb.sheetnames) == 1
    ws = wb.active
    assert ws.title == '19.08. – 01.09.2026'
    assert len(ws.title) <= 31


def test_title_and_client_rows_merged_and_styled():
    ws = _load().active
    assert 'A1:C1' in [str(r) for r in ws.merged_cells.ranges]
    assert 'A2:C2' in [str(r) for r in ws.merged_cells.ranges]
    a1 = ws['A1']
    assert a1.value == 'Výkaz práce – 19.08.2026 → 01.09.2026'
    assert a1.font.bold is True
    assert a1.font.size == 14
    assert a1.font.color.rgb == 'FFFFFFFF'
    assert a1.fill.fgColor.rgb == 'FF1F4E78'
    assert a1.alignment.horizontal == 'center'
    a2 = ws['A2']
    assert a2.value == 'Klient: Testovací klient'
    assert a2.font.size == 10


def test_row3_empty_and_header_row4():
    ws = _load().active
    assert ws['A3'].value is None
    assert ws['B3'].value is None
    assert ws['C3'].value is None
    assert [ws['A4'].value, ws['B4'].value, ws['C4'].value] == ['Dátum', 'Hodiny', 'Popis činnosti']
    for coord in ('A4', 'B4', 'C4'):
        c = ws[coord]
        assert c.font.bold is True
        assert c.font.color.rgb == 'FFFFFFFF'
        assert c.fill.fgColor.rgb == 'FF2E75B6'
        assert c.alignment.horizontal == 'center'


def test_data_rows_values_formats_and_alignment():
    ws = _load().active
    # riadky 5,6,7 pre 3 položky
    assert _as_date(ws['A5'].value) == dt.date(2026, 8, 19)
    assert ws['B5'].value == 3.0
    assert ws['C5'].value == 'návrh výkresu'
    assert _as_date(ws['A7'].value) == dt.date(2026, 8, 25)
    assert ws['B7'].value == 1.5
    # number formaty
    assert ws['A5'].number_format == r'yyyy\-mm\-dd'
    assert ws['B5'].number_format == '0.0'
    # zarovnanie
    assert ws['A5'].alignment.horizontal == 'center'
    assert ws['B5'].alignment.horizontal == 'center'
    assert ws['C5'].alignment.horizontal in (None, 'left', 'general')
    # veľkosť písma dátových buniek
    assert ws['A5'].font.size == 10
    assert ws['C5'].font.size == 10


def test_zebra_every_second_data_row():
    ws = _load().active
    # data riadok 0 (r5) bez výplne, riadok 1 (r6) F2F2F2, riadok 2 (r7) bez výplne
    assert ws['A5'].fill.patternType in (None, 'none')
    assert ws['A6'].fill.fgColor.rgb == 'FFF2F2F2'
    assert ws['B6'].fill.fgColor.rgb == 'FFF2F2F2'
    assert ws['C6'].fill.fgColor.rgb == 'FFF2F2F2'
    assert ws['A7'].fill.patternType in (None, 'none')


def test_total_row_formula_and_style():
    ws = _load().active
    # 3 položky → dáta 5..7, SPOLU v riadku 8
    assert ws['A8'].value == 'SPOLU'
    assert ws['B8'].value == '=SUM(B5:B7)'
    assert ws['C8'].value == 'hodín'
    assert ws['B8'].number_format == '0.0'
    for coord in ('A8', 'B8', 'C8'):
        c = ws[coord]
        assert c.font.bold is True
        assert c.fill.fgColor.rgb == 'FFDDEBF7'


def test_column_widths():
    ws = _load().active
    assert ws.column_dimensions['A'].width == 14
    assert ws.column_dimensions['B'].width == 10
    assert ws.column_dimensions['C'].width == 95


def test_no_person_name_in_any_cell():
    ws = _load().active
    texts = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    blob = '\n'.join(texts)
    assert 'Ján Novák' not in blob
    assert 'Peter Kováč' not in blob


def test_client_none_omits_row2_but_keeps_layout():
    ws = _load(client=None).active
    assert ws['A2'].value is None
    assert 'A2:C2' not in [str(r) for r in ws.merged_cells.ranges]
    # hlavička ostáva v riadku 4, dáta od 5, SPOLU v 8
    assert ws['A4'].value == 'Dátum'
    assert ws['B8'].value == '=SUM(B5:B7)'


def test_empty_items_no_crash_total_zero():
    ws = _load(items=[]).active
    assert ws['A4'].value == 'Dátum'
    # SPOLU hneď pod hlavičkou (riadok 5), bez self-referenčného SUM
    assert ws['A5'].value == 'SPOLU'
    assert ws['B5'].value in (0, 0.0)
    assert ws['C5'].value == 'hodín'


def test_filename_ascii_safe():
    assert xlsx_filename(OD, DO, 'Testovací klient') == 'Vykaz_prace_2026-08-19_2026-09-01_Testovaci_klient.xlsx'
    assert xlsx_filename(OD, DO, None) == 'Vykaz_prace_2026-08-19_2026-09-01.xlsx'
    assert xlsx_filename(OD, DO, '') == 'Vykaz_prace_2026-08-19_2026-09-01.xlsx'


def test_sheet_name_strips_forbidden_chars_and_truncates():
    # priamy test sanitizéra — každý zakázaný znak ([ ] : * ? / \) → medzera,
    # výsledok vždy <= 31 znakov
    raw = 'a[b]c:d*e?f/g\\h' + 'x' * 40
    out = _sanitize_sheet_name(raw)
    assert not (set(out) & set('[]:*?/\\'))
    assert len(out) <= 31
    # bežný titulok nášho formátu prejde nedotknutý (len orez apostrofov)
    ws = _load().active
    assert not (set(ws.title) & set('[]:*?/\\'))


def test_desc_starting_with_equals_is_text_not_formula():
    # popis zo správy kanála začínajúci '=' NESMIE ožiť ako vzorec v XLSX
    items = [(dt.datetime(2026, 8, 19, 10, 0, tzinfo=TZ), 'X', 2.0,
              '=HYPERLINK("http://zlo";"klik")')]
    ws = _load(items=items).active
    c = ws['C5']
    assert c.value == '=HYPERLINK("http://zlo";"klik")'
    assert c.data_type == 's'  # text, nie 'f' (formula)
